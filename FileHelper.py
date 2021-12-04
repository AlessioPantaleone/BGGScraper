import csv
import re
from openpyxl import Workbook


def get_input_from_file(filename: str):
    with open(filename, 'r') as fd:
        reader = csv.reader(fd)
        data = []
        for row in reader:
            data.append(re.findall(r'\d+', str(row))[0])
    return data


def export_data_to_excel(data, filename: str):
    workbook = Workbook()

    """   OLD EXPORT METHOD
    sheet = workbook.create_sheet("Da aggiungere")
    for i, array in enumerate(data):
        for j, cell in enumerate(array):
            if cell == "Mechanics" or cell == "Designers" or cell == "Artists" or cell == "Publishers":
                new_value = ""
                for name in array[cell]:
                    new_value += f"{name}\n"
                sheet.cell(row=i + 1, column=j + 1).value = new_value
            else:
                sheet.cell(row=i + 1, column=j + 1).value = str(array[cell])
    """

    sheet = workbook.create_sheet("Da aggiungere")

    for i, array in enumerate(data):
        sheet.cell(row=i + 1, column=1).value = array["Name"]
        sheet.cell(row=i + 1, column=2).value = f"https://boardgamegeek.com/boardgame/{array['id']}"
        sheet.cell(row=i + 1, column=6).value = array["PublicationYear"]
        sheet.cell(row=i + 1, column=7).value = "Playadice"
        sheet.cell(row=i + 1, column=8).value = "Brucalab"
        sheet.cell(row=i + 1, column=9).value = "TODO"
        sheet.cell(row=i + 1, column=10).value = array["BGG_Score"]
        sheet.cell(row=i + 1, column=11).value = array["MinPlayers"]
        sheet.cell(row=i + 1, column=12).value = array["MaxPlayers"]
        sheet.cell(row=i + 1, column=13).value = array["MinAge"]
        sheet.cell(row=i + 1, column=14).value = array["PlayingTime"]

        gamecategory = "Not Categorized"
        if array["Category"] == "familygames":
            gamecategory = "Famiglia"
        if array["Category"] == "strategygames":
            gamecategory = "Strategia"
        if array["Category"] == "partygames":
            gamecategory = "Party"
        if array["Category"] == "abstracts":
            gamecategory = "Astratto"
        if array["Category"] == "childrensgames":
            gamecategory = "Famiglia"
        if array["Category"] == "cgs":
            gamecategory = "Famiglia"
        if array["Category"] == "wargames":
            gamecategory = "Guerra"
        if array["Category"] == "thematic":
            gamecategory = "Tematico"
        sheet.cell(row=i + 1, column=15).value = gamecategory

        new_value = ""
        for j, name in enumerate(array["Mechanics"]):
            if j < 3:
                new_value += f"{name}\n"
        sheet.cell(row=i + 1, column=16).value = new_value

        new_value = ""
        for j, name in enumerate(array["Designers"]):
            if j < 3:
                if name != "(Uncredited)":
                    new_value += f"{name}\n"
        sheet.cell(row=i + 1, column=17).value = new_value

        new_value = ""
        for j, name in enumerate(array["Artists"]):
            if j < 3:
                if name != "(Uncredited)":
                    new_value += f"{name}\n"
        sheet.cell(row=i + 1, column=18).value = new_value

        new_value = ""
        for j, name in enumerate(array["Publishers"]):
            if j < 3:
                if name != "(Public Domain)":
                    new_value += f"{name}\n"
        sheet.cell(row=i + 1, column=19).value = new_value

    workbook.save(filename=filename)
